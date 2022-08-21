#!/usr/bin/env python3

# Python 3.9.5

# 01_using_openpyxl.py

# Dependencies
import openpyxl
import os

path = 'C:\\Users\\user\\spreadsheets'
os.chdir(path)

for dirname, subfolders, filenames in os.walk(path):
    for filename in filenames:
        if filename.find('xls') != -1:
            workbook = openpyxl.load_workbook(filename, data_only=True)

# Access individual Worksheet.
worksheet = workbook["Sheet1"]
worksheet = workbook.worksheets[0]

# Return a list of all WorkSheet names.
workbook.sheetnames

# Get the number of WorkSheets in the WorkBook.
number_of_sheets = len(workbook.sheetnames)

# Alternatively loop through the list:
for sheetname in workbook.sheetnames:
    print(sheetname)

# Return used range of the WorkSheet:
worksheet.max_row, worksheet.max_column

# Read a single cell's value
worksheet["A1"].value
