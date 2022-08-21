#!/usr/bin/env python3

# Python 3.9.5

# 04_write_with_openpyxl.py

# Dependencies
import datetime as dt
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill

import os

path = 'C:\\Users\\user\\spreadsheets'
os.chdir(path)

# Initiate a WorkBook
workbook = openpyxl.Workbook()

# Activate the first WorkSheet
worksheet = workbook.active
worksheet.title = "Sheet1"

# Write to individual cells using different notations
worksheet["A1"].value = "A1-Notation"
worksheet.cell(row=2, column=1, value="Row-Cell-Notation")

# Formatting: color, alignment, border, font
font_format = Font(color="FF0000", bold=True)
thin = Side(border_style="thin", color="FF0000")

worksheet["A3"].value = "Formatted text"
worksheet["A3"].font = font_format
worksheet["A3"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
worksheet["A3"].alignment = Alignment(horizontal="center")
worksheet["A3"].fill = PatternFill(fgColor="00FF00", fill_type="solid")

# Number formatting
worksheet["A4"].value = 1.23456789
worksheet["A4"].number_format = "0.000"

# Date formatting
worksheet["A5"].value = dt.date(2022, 8, 21)
worksheet["A5"].number_format = "dd/mm/yyyy"

# Using formula (comma as delimiter)
worksheet["A6"].value = "=SUM(A4, 2)"

# Adding images, previously set a path to your image directory
worksheet.add_image(Image("img/image.png"), "D1")

# Adding data
worksheet["B10"].value = "Division 12"
worksheet["C10"].value = "Division 19"
worksheet["A11"].value = 2021
worksheet["A12"].value = 2022
worksheet["B11"].value = 225
worksheet["C11"].value = 495
worksheet["B12"].value = 450
worksheet["C12"].value = 700

# Adding chart
chart = BarChart()
chart.type = "col"
chart.title = "Summary"
chart.x_axis.title = "Region"
chart.y_axis.title = "Output"
chart_data = Reference(worksheet, min_row=11, min_col=1, max_row=12, max_col=3)
chart_categories = Reference(worksheet, min_row=10, min_col=2, max_row=10, max_col=3)

chart.add_data(chart_data, titles_from_data=True, from_rows=True)
chart.set_categories(chart_categories)
worksheet.add_chart(chart, "A15")

# Save WorkBook, create a file on your hard drive
workbook.save("ThisWorkBook.xlsx")
